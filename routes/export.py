import csv
import json
import os
from datetime import datetime
from html import escape

from fastapi import APIRouter, Body, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from services.dataset_service import REPORT_DIR, ensure_data_dirs
from services.dictionary_cache import get_result, update_result


router = APIRouter()

# Studio palette (unchanged)
COLOR_BG = "#f7f3ea"
COLOR_INK = "#111214"
COLOR_CREAM = "#fffaf0"
COLOR_MUTED = "#6b665e"
COLOR_BORDER = "#d8d0c2"
COLOR_TEAL = "#00a8c5"


def _safe_filename(name: str) -> str:
    return os.path.basename(name or "dataset").replace(" ", "_")


def _fairness_text(flag: dict | None) -> str:
    if not flag:
        return ""
    groq = flag.get("groq_verification") or {}
    verification = ""
    if groq:
        verification = f" Groq: {groq.get('verdict', 'Uncertain')} — {groq.get('reason', '')}"
    return f"{flag.get('eu_ai_act_article') or ''} {flag.get('reason') or ''}{verification}".strip()


def _p(text) -> str:
    return escape(str(text or ""))


def _truncate(text: str, limit: int = 1200) -> str:
    text = str(text or "").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _draw_watermark(canvas, doc) -> None:
    canvas.saveState()
    width, height = letter
    canvas.setFillColor(colors.HexColor(COLOR_BG))
    canvas.rect(0, 0, width, height, fill=1, stroke=0)
    canvas.setStrokeColor(colors.HexColor("#e6ded0"))
    canvas.setLineWidth(0.6)
    for x in range(36, int(width), 72):
        canvas.line(x, 0, x, height)
    for y in range(36, int(height), 72):
        canvas.line(0, y, width, y)

    canvas.translate(width / 2, height / 2)
    canvas.rotate(32)
    canvas.setFillColor(colors.Color(0.06, 0.06, 0.07, alpha=0.045))
    canvas.setFont("Helvetica-Bold", 76)
    canvas.drawCentredString(0, 0, "DataLens")
    canvas.setFillColor(colors.Color(0.0, 0.65, 0.78, alpha=0.08))
    canvas.setFont("Helvetica-Bold", 20)
    canvas.drawCentredString(0, -30, "DATA DICTIONARY")
    canvas.restoreState()

    canvas.saveState()
    canvas.setFillColor(colors.HexColor(COLOR_INK))
    canvas.roundRect(36, height - 42, 24, 24, 6, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor(COLOR_CREAM))
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawCentredString(48, height - 27, "DL")
    canvas.setFillColor(colors.HexColor("#4c473f"))
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(66, height - 28, "DataLens Studio")
    canvas.setFillColor(colors.HexColor("#777166"))
    canvas.drawRightString(width - 36, 22, f"Page {doc.page}")
    canvas.restoreState()


def _pdf_styles():
    base = getSampleStyleSheet()
    base.add(ParagraphStyle(
        name="StudioTitle",
        parent=base["Title"],
        fontName="Helvetica-Bold",
        fontSize=26,
        leading=30,
        textColor=colors.HexColor(COLOR_INK),
        spaceAfter=10,
    ))
    base.add(ParagraphStyle(
        name="StudioH2",
        parent=base["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        textColor=colors.HexColor(COLOR_INK),
        spaceBefore=14,
        spaceAfter=8,
    ))
    base.add(ParagraphStyle(
        name="StudioH3",
        parent=base["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor(COLOR_INK),
        spaceBefore=6,
        spaceAfter=4,
    ))
    base.add(ParagraphStyle(
        name="StudioBody",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#332f29"),
        alignment=TA_LEFT,
        spaceAfter=6,
    ))
    base.add(ParagraphStyle(
        name="StudioMuted",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor(COLOR_MUTED),
        spaceAfter=4,
    ))
    base.add(ParagraphStyle(
        name="StudioCode",
        parent=base["Code"],
        fontName="Courier",
        fontSize=7.5,
        leading=10,
        textColor=colors.HexColor(COLOR_CREAM),
        backColor=colors.HexColor(COLOR_INK),
        borderPadding=6,
        leftIndent=4,
        rightIndent=4,
        spaceAfter=6,
    ))
    base.add(ParagraphStyle(
        name="StudioCell",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#332f29"),
    ))
    return base


def _kpi_table(rows: list[list], col_widths: list[float]) -> Table:
    table = Table(rows, colWidths=col_widths, repeatRows=0)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_INK)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(COLOR_CREAM)),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(COLOR_CREAM)),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor(COLOR_INK)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 16),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor(COLOR_INK)),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor(COLOR_BORDER)),
    ]))
    return table


def _data_table(headers: list, rows: list, col_widths: list, styles) -> Table:
    body = [headers] + rows
    table = Table(body, colWidths=col_widths, repeatRows=1, splitByRow=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_INK)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(COLOR_CREAM)),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(COLOR_CREAM)),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(COLOR_BORDER)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return table


def _column_pdf_block(profile: dict, styles, content_width: float) -> list:
    """Flowable block per column — avoids table SPAN layout that blows up page height."""
    flow = []
    name = profile.get("display_name") or profile.get("column_name")
    semantic = profile.get("semantic_type") or {}
    sem_label = semantic.get("label", "")
    sem_conf = int(float(semantic.get("confidence", 0)) * 100) if semantic.get("confidence") else ""

    flow.append(Paragraph(f"<b>{_p(name)}</b>", styles["StudioH3"]))
    meta_bits = [
        f"<b>Type:</b> {_p(profile.get('technical_type') or profile.get('dtype'))}",
        f"<b>Null:</b> {_p(profile.get('null_pct'))}%",
        f"<b>Unique:</b> {_p(profile.get('unique_count'))}",
    ]
    if sem_label:
        meta_bits.append(f"<b>Semantic:</b> {_p(sem_label)} ({sem_conf}%)")
    flow.append(Paragraph(" &nbsp;|&nbsp; ".join(meta_bits), styles["StudioMuted"]))

    desc = _truncate(profile.get("business_description") or profile.get("description") or "", 900)
    if desc:
        flow.append(Paragraph(_p(desc), styles["StudioBody"]))

    notes = profile.get("quality_notes") or []
    if notes:
        flow.append(Paragraph(
            "<b>Quality:</b> " + _p("; ".join(str(n) for n in notes[:3])),
            styles["StudioMuted"],
        ))

    if profile.get("anomaly_note"):
        flow.append(Paragraph(f"<b>Anomaly:</b> {_p(_truncate(profile.get('anomaly_note'), 400))}", styles["StudioMuted"]))

    flag = profile.get("fairness_flag")
    if flag:
        flow.append(Paragraph(f"<b>Fairness:</b> {_p(_truncate(_fairness_text(flag), 500))}", styles["StudioMuted"]))

    flow.append(Spacer(1, 8))
    return flow


def create_dictionary_pdf(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_dictionary.pdf")
    doc = SimpleDocTemplate(
        path,
        pagesize=letter,
        rightMargin=0.72 * inch,
        leftMargin=0.72 * inch,
        topMargin=0.85 * inch,
        bottomMargin=0.72 * inch,
    )
    content_width = doc.width
    styles = _pdf_styles()
    story = []

    metadata = result.get("metadata", {})
    readiness = result.get("readiness", {})
    health = result.get("health", {})
    trust = health.get("overall_trust") or {}
    dataset_story = result.get("story") or {}
    quality_audit = result.get("quality_audit") or {}

    story.append(Paragraph("DataLens Data Dictionary", styles["StudioTitle"]))
    story.append(Paragraph(
        f"<b>{_p(metadata.get('filename', 'Unknown'))}</b> · "
        f"{_p(metadata.get('row_count', 0))} rows · {_p(metadata.get('column_count', 0))} columns · "
        f"Generated {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}",
        styles["StudioMuted"],
    ))
    story.append(Spacer(1, 12))

    quarter = content_width / 4
    story.append(_kpi_table(
        [
            ["Rows", "Columns", "Trust", "AI readiness"],
            [
                str(metadata.get("row_count", 0)),
                str(metadata.get("column_count", 0)),
                str(trust.get("score", quality_audit.get("score", "—"))),
                f"{readiness.get('score', '—')} ({readiness.get('grade', 'N/A')})",
            ],
        ],
        [quarter] * 4,
    ))
    story.append(Spacer(1, 12))

    if dataset_story.get("executive_summary"):
        story.append(Paragraph("Executive summary", styles["StudioH2"]))
        story.append(Paragraph(_p(_truncate(dataset_story["executive_summary"], 1500)), styles["StudioBody"]))

    if trust.get("score") is not None:
        story.append(Paragraph(
            f"<b>Overall trust:</b> {_p(trust.get('score'))}/100 — {_p(_truncate(trust.get('reasoning', ''), 400))}",
            styles["StudioBody"],
        ))

    health_rows = []
    for key in ("completeness", "consistency", "validity", "uniqueness", "governance", "documentation_quality"):
        dim = health.get(key) or {}
        if dim.get("score") is not None:
            health_rows.append([
                key.replace("_", " ").title(),
                str(dim.get("score")),
                Paragraph(_p(_truncate(dim.get("reasoning", ""), 200)), styles["StudioCell"]),
            ])
    if health_rows:
        story.append(Paragraph("Trust dimensions", styles["StudioH2"]))
        hw = [content_width * 0.22, content_width * 0.12, content_width * 0.66]
        story.append(_data_table(
            ["Dimension", "Score", "Summary"],
            health_rows,
            hw,
            styles,
        ))
        story.append(Spacer(1, 10))

    if quality_audit:
        story.append(Paragraph("Quality audit", styles["StudioH2"]))
        story.append(Paragraph(
            f"Health score <b>{_p(quality_audit.get('score', 0))}/100</b> · "
            f"Critical/high: <b>{_p(quality_audit.get('critical_count', 0))}</b> · "
            f"Warnings: <b>{_p(quality_audit.get('warning_count', 0))}</b> · "
            f"Suggestions: <b>{_p(quality_audit.get('suggestion_count', 0))}</b>",
            styles["StudioBody"],
        ))
        finding_rows = []
        for finding in (quality_audit.get("top_findings") or quality_audit.get("findings") or [])[:12]:
            finding_rows.append([
                _p(finding.get("severity")),
                Paragraph(
                    f"<b>{_p(finding.get('title'))}</b><br/>{_p(_truncate(finding.get('summary', ''), 280))}",
                    styles["StudioCell"],
                ),
                Paragraph(_p(_truncate(finding.get("recommendation", ""), 220)), styles["StudioCell"]),
            ])
        if finding_rows:
            fw = [content_width * 0.14, content_width * 0.46, content_width * 0.40]
            story.append(_data_table(["Severity", "Finding", "Action"], finding_rows, fw, styles))
        story.append(Spacer(1, 8))

    dictionary = result.get("column_dictionary") or result.get("profiles", [])
    if dictionary:
        story.append(PageBreak())
        story.append(Paragraph("Data dictionary", styles["StudioH2"]))
        story.append(Paragraph(
            f"{len(dictionary)} columns documented below.",
            styles["StudioMuted"],
        ))
        story.append(Spacer(1, 6))
        for profile in dictionary:
            story.extend(_column_pdf_block(profile, styles, content_width))

    relationships = result.get("relationships") or []
    redundant = result.get("redundant_columns") or []
    if relationships or redundant:
        story.append(Paragraph("Relationships", styles["StudioH2"]))
        rel_rows = []
        for rel in relationships[:20]:
            rel_rows.append([
                _p(rel.get("col_a")),
                _p(rel.get("col_b")),
                _p(rel.get("correlation")),
                _p(rel.get("type")),
                Paragraph(_p(_truncate(rel.get("note", ""), 300)), styles["StudioCell"]),
            ])
        for rel in redundant[:10]:
            rel_rows.append([
                _p(rel.get("col_a")),
                _p(rel.get("col_b")),
                f"{_p(rel.get('match_pct'))}%",
                "Near duplicate",
                Paragraph(_p(_truncate(rel.get("note", ""), 300)), styles["StudioCell"]),
            ])
        if rel_rows:
            rw = [
                content_width * 0.18,
                content_width * 0.18,
                content_width * 0.12,
                content_width * 0.18,
                content_width * 0.34,
            ]
            story.append(_data_table(
                ["Column A", "Column B", "Strength", "Type", "Note"],
                rel_rows,
                rw,
                styles,
            ))

    queries = result.get("query_suggestions") or []
    if queries:
        story.append(Paragraph("Suggested analyses", styles["StudioH2"]))
        for query in queries[:8]:
            story.append(Paragraph(f"<b>{_p(query.get('question', ''))}</b>", styles["StudioBody"]))
            if query.get("pandas_query"):
                story.append(Paragraph(f"Pandas: {_p(query.get('pandas_query'))}", styles["StudioCode"]))
            if query.get("sql_query"):
                story.append(Paragraph(f"SQL: {_p(query.get('sql_query'))}", styles["StudioCode"]))
            story.append(Spacer(1, 6))

    doc.build(story, onFirstPage=_draw_watermark, onLaterPages=_draw_watermark)
    return path


def _excel_header_style():
    return (
        PatternFill("solid", fgColor="111214"),
        Font(name="Calibri", bold=True, color="FFFAF0", size=10),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
    )


def _excel_border():
    thin = Side(style="thin", color="D8D0C2")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header_row(ws, row_num: int = 1) -> None:
    fill, font, align = _excel_header_style()
    border = _excel_border()
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[row_num].height = 26


def _apply_row_style(ws, row_num: int, fill: PatternFill, center_cols: set[int] | None = None) -> None:
    border = _excel_border()
    center_cols = center_cols or set()
    for cell in ws[row_num]:
        cell.fill = fill
        cell.border = border
        if cell.column in center_cols:
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    text_len = max((len(str(cell.value or "")) for cell in ws[row_num]), default=0)
    ws.row_dimensions[row_num].height = max(28, min(80, 22 + text_len // 90 * 10))


def create_dictionary_excel(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_dictionary.xlsx")
    wb = Workbook()

    red = PatternFill("solid", fgColor="FAD7D7")
    amber = PatternFill("solid", fgColor="FBE6B5")
    green = PatternFill("solid", fgColor="E7F3C5")
    cream = PatternFill("solid", fgColor="FFFAF0")

    metadata = result.get("metadata", {})
    health = result.get("health", {})
    story = result.get("story", {})
    readiness = result.get("readiness", {})

    # --- Overview sheet ---
    overview = wb.active
    overview.title = "Overview"
    overview.column_dimensions["A"].width = 22
    overview.column_dimensions["B"].width = 72
    rows = [
        ("DataLens export", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Dataset", metadata.get("filename")),
        ("Rows", metadata.get("row_count")),
        ("Columns", metadata.get("column_count")),
        ("Overall trust", (health.get("overall_trust") or {}).get("score")),
        ("AI & Analytics Readiness", f"{readiness.get('score')} (grade {readiness.get('grade')})"),
        ("", ""),
        ("Executive summary", story.get("executive_summary", "")),
        ("Business summary", story.get("business_summary", "")),
    ]
    for label, value in rows:
        overview.append([label, value])
    for row in overview.iter_rows(min_row=1, max_row=overview.max_row):
        row[0].font = Font(name="Calibri", bold=True, size=10)
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].value in {"Executive summary", "Business summary"}:
            overview.row_dimensions[row[0].row].height = 48

    # --- Data dictionary sheet ---
    ws = wb.create_sheet("Data Dictionary")
    headers = [
        "Column", "Display name", "Technical type", "Semantic type", "Null %",
        "Unique", "Confidence", "Sensitivity", "Description", "Quality notes",
        "Anomaly", "Fairness", "Groq verdict",
    ]
    ws.append(headers)
    _style_header_row(ws)
    center_cols = {5, 6, 7, 8, 13}

    dictionary = result.get("column_dictionary") or result.get("profiles", [])
    for profile in dictionary:
        semantic = profile.get("semantic_type") or {}
        sensitivity = profile.get("sensitivity") or {}
        flag = profile.get("fairness_flag") or {}
        groq = flag.get("groq_verification") or {}
        notes = "; ".join(profile.get("quality_notes") or [])
        ws.append([
            profile.get("column_name"),
            profile.get("display_name"),
            profile.get("technical_type") or profile.get("dtype"),
            semantic.get("label", ""),
            profile.get("null_pct"),
            profile.get("unique_count"),
            profile.get("confidence"),
            sensitivity.get("level", "Low"),
            profile.get("business_description") or profile.get("description"),
            notes,
            profile.get("anomaly_note"),
            _fairness_text(flag) if flag else "",
            groq.get("verdict", ""),
        ])
        row = ws.max_row
        if flag:
            fill = red
        elif profile.get("anomaly_note"):
            fill = amber
        elif profile.get("confidence") in ("Confirmed", 0.95) or (
            isinstance(profile.get("confidence"), (int, float)) and float(profile.get("confidence", 0)) >= 0.9
        ):
            fill = green
        else:
            fill = cream
        _apply_row_style(ws, row, fill, center_cols)

    col_widths = [16, 18, 14, 18, 8, 10, 10, 12, 48, 32, 28, 36, 14]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    # --- Audit sheet ---
    audit = result.get("quality_audit") or {}
    if audit.get("findings"):
        audit_ws = wb.create_sheet("Quality Audit")
        audit_ws.append(["Metric", "Value"])
        audit_ws.append(["Health score", audit.get("score")])
        audit_ws.append(["Critical / high", audit.get("critical_count")])
        audit_ws.append(["Warnings", audit.get("warning_count")])
        audit_ws.append(["Suggestions", audit.get("suggestion_count")])
        audit_ws.append([])
        audit_ws.append(["Severity", "Category", "Title", "Summary", "Columns", "Recommendation"])
        _style_header_row(audit_ws, audit_ws.max_row)
        for finding in audit.get("findings", []):
            audit_ws.append([
                finding.get("severity"),
                finding.get("category"),
                finding.get("title"),
                finding.get("summary"),
                "; ".join(finding.get("source_columns") or []),
                finding.get("recommendation"),
            ])
            row = audit_ws.max_row
            sev = finding.get("severity")
            fill = red if sev in {"Critical", "High"} else amber if sev == "Medium" else green
            _apply_row_style(audit_ws, row, fill, {1, 2})
        for letter, width in zip("ABCDEF", [12, 16, 28, 44, 20, 44]):
            audit_ws.column_dimensions[letter].width = width

    # --- Governance sheet ---
    gov = result.get("governance") or {}
    gov_ws = wb.create_sheet("Governance")
    gov_ws.column_dimensions["A"].width = 24
    gov_ws.column_dimensions["B"].width = 68
    gov_rows = [
        ("Risk level", gov.get("risk_level")),
        ("PII fields", ", ".join(gov.get("detected_pii") or [])),
        ("Flagged columns", gov.get("flagged_column_count")),
    ]
    for item in gov.get("recommendations") or []:
        gov_rows.append(("Recommendation", item))
    for item in gov.get("gdpr_notes") or []:
        gov_rows.append(("GDPR", item))
    for label, value in gov_rows:
        gov_ws.append([label, value])

    wb.save(path)
    return path


def create_dictionary_markdown(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_dictionary.md")
    metadata = result.get("metadata", {})
    story = result.get("story", {})
    health = result.get("health", {})
    readiness = result.get("readiness", {})
    gov = result.get("governance", {})
    audit = result.get("quality_audit", {})

    lines = [
        f"# DataLens Data Dictionary",
        "",
        f"**Dataset:** {metadata.get('filename', 'unknown')}  ",
        f"**Rows:** {metadata.get('row_count', 0):,} · **Columns:** {metadata.get('column_count', 0)}  ",
        f"**Exported:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        "",
        "---",
        "",
        "## Executive summary",
        "",
        story.get("executive_summary", "_No summary available._"),
        "",
        "## Trust & readiness",
        "",
        f"| Metric | Score |",
        f"|--------|-------|",
        f"| Overall trust | {(health.get('overall_trust') or {}).get('score', '—')} |",
        f"| AI & Analytics Readiness | {readiness.get('score', '—')} ({readiness.get('grade', '')}) |",
        "",
    ]
    for key in ("completeness", "consistency", "validity", "uniqueness", "governance", "documentation_quality"):
        dim = health.get(key) or {}
        if dim.get("score") is not None:
            lines.append(f"- **{key.replace('_', ' ').title()}:** {dim.get('score')}/100 — {dim.get('reasoning', '')}")
    lines.extend(["", "## Governance", ""])
    lines.append(f"- **Risk level:** {gov.get('risk_level', 'Unknown')}")
    lines.append(f"- **PII fields:** {', '.join(gov.get('detected_pii') or []) or 'None'}")
    for rec in gov.get("recommendations") or []:
        lines.append(f"- {rec}")

    if audit.get("findings"):
        lines.extend(["", "## Quality audit", "", f"**Score:** {audit.get('score')}/100", ""])
        for finding in audit.get("findings", [])[:15]:
            lines.append(f"### {finding.get('severity')} — {finding.get('title')}")
            lines.append(finding.get("summary", ""))
            cols = ", ".join(finding.get("source_columns") or [])
            if cols:
                lines.append(f"*Columns:* {cols}")
            lines.append("")

    lines.extend(["", "## Column dictionary", ""])
    for entry in result.get("column_dictionary") or result.get("profiles", []):
        semantic = entry.get("semantic_type") or {}
        lines.extend([
            f"### {entry.get('display_name') or entry.get('column_name')}",
            "",
            f"| Attribute | Value |",
            f"|-----------|-------|",
            f"| Technical type | `{entry.get('technical_type') or entry.get('dtype')}` |",
            f"| Semantic type | {semantic.get('label', '—')} ({int(float(semantic.get('confidence', 0)) * 100)}%) |",
            f"| Null % | {entry.get('null_pct')} |",
            f"| Unique values | {entry.get('unique_count')} |",
            f"| Sensitivity | {(entry.get('sensitivity') or {}).get('level', 'Low')} |",
            "",
            entry.get("business_description") or entry.get("description") or "_No description._",
            "",
        ])

    with open(path, "w", encoding="utf-8") as md_file:
        md_file.write("\n".join(lines))
    return path


def create_audit_csv(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_audit.csv")
    audit = result.get("quality_audit") or {}
    metadata = result.get("metadata", {})
    with open(path, "w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["DataLens quality audit export"])
        writer.writerow(["dataset", metadata.get("filename", "")])
        writer.writerow(["health_score", audit.get("score", "")])
        writer.writerow(["exported_at", datetime.utcnow().isoformat()])
        writer.writerow([])
        writer.writerow([
            "severity", "category", "title", "summary",
            "source_columns", "confidence", "affected_rows", "recommendation",
        ])
        for finding in audit.get("findings", []):
            writer.writerow([
                finding.get("severity"),
                finding.get("category"),
                finding.get("title"),
                finding.get("summary"),
                "; ".join(finding.get("source_columns") or []),
                finding.get("confidence"),
                finding.get("affected_rows"),
                finding.get("recommendation"),
            ])
    return path


def create_governance_report(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_governance.json")
    governance = result.get("governance") or {}
    payload = {
        "report_type": "DataLens Governance & Compliance",
        "generated_at": datetime.utcnow().isoformat(),
        "dataset": {
            "filename": result.get("metadata", {}).get("filename"),
            "row_count": result.get("metadata", {}).get("row_count"),
            "column_count": result.get("metadata", {}).get("column_count"),
        },
        "governance": governance,
        "health_scores": {
            key: value for key, value in (result.get("health") or {}).items()
            if isinstance(value, dict) and "score" in value
        },
        "fairness_columns": [
            {
                "column": entry.get("column_name"),
                "fairness_flag": entry.get("fairness_flag"),
            }
            for entry in (result.get("column_dictionary") or [])
            if entry.get("fairness_flag")
        ],
    }
    with open(path, "w", encoding="utf-8") as report_file:
        json.dump(payload, report_file, indent=2, default=str)
    return path


def create_dictionary_json(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_dictionary.json")
    export_payload = {
        "exported_at": datetime.utcnow().isoformat(),
        "export_version": 2,
        "analysis": result,
    }
    with open(path, "w", encoding="utf-8") as json_file:
        json.dump(export_payload, json_file, indent=2, default=str)
    return path


def export_result(session_id: str, export_format: str, result: dict) -> FileResponse:
    export_format = export_format.lower()
    base = _safe_filename(result.get("metadata", {}).get("filename", "dataset")).replace(".csv", "").replace(".xlsx", "")
    filenames = {
        "pdf": f"datalens_{base}_dictionary.pdf",
        "excel": f"datalens_{base}_dictionary.xlsx",
        "json": f"datalens_{base}_package.json",
        "markdown": f"datalens_{base}_dictionary.md",
        "audit_csv": f"datalens_{base}_audit.csv",
        "governance": f"datalens_{base}_governance.json",
    }
    if export_format == "pdf":
        path = create_dictionary_pdf(session_id, result)
        return FileResponse(path, media_type="application/pdf", filename=filenames["pdf"])
    if export_format == "excel":
        path = create_dictionary_excel(session_id, result)
        return FileResponse(
            path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filenames["excel"],
        )
    if export_format == "json":
        path = create_dictionary_json(session_id, result)
        return FileResponse(path, media_type="application/json", filename=filenames["json"])
    if export_format == "markdown":
        path = create_dictionary_markdown(session_id, result)
        return FileResponse(path, media_type="text/markdown", filename=filenames["markdown"])
    if export_format in {"audit_csv", "audit-csv", "csv"}:
        path = create_audit_csv(session_id, result)
        return FileResponse(path, media_type="text/csv", filename=filenames["audit_csv"])
    if export_format == "governance":
        path = create_governance_report(session_id, result)
        return FileResponse(path, media_type="application/json", filename=filenames["governance"])
    raise HTTPException(
        status_code=400,
        detail="Format must be pdf, excel, json, markdown, audit_csv, or governance",
    )


@router.get("/export/{session_id}/{format}")
async def export_cached_result(session_id: str, format: str):
    result = get_result(session_id)
    if not result:
        raise HTTPException(status_code=404, detail="Session result not found")
    return export_result(session_id, format, result)


@router.post("/export/{session_id}/{format}")
async def export_edited_result(session_id: str, format: str, result: dict = Body(...)):
    update_result(session_id, result)
    return export_result(session_id, format, result)
